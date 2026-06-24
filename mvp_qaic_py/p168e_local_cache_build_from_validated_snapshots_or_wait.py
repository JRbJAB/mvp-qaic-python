"""P168E local cache build from validated snapshots or wait.

This module stays inside MVP_QAIC_PY. It consumes P168D local operator export
validation outputs and prepares a bounded local cache package only when validated
operator snapshot files are present. It does not call Google APIs, does not write
Google Sheets, does not execute Apps Script, and does not touch broker/order flows.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

STEP = "P168E_LOCAL_CACHE_BUILD_FROM_VALIDATED_SNAPSHOTS_OR_WAIT"
STATUS_WAIT = "P168E_LOCAL_CACHE_BUILD_WAIT_VALIDATED_SNAPSHOTS_READONLY"
STATUS_READY = "P168E_LOCAL_CACHE_BUILD_READY_FROM_VALIDATED_SNAPSHOTS_READONLY"
EXPORT_PREFIX_P168D = "P168D_LOCAL_SNAPSHOT_IMPORT_READER_FROM_OPERATOR_EXPORTS_"
EXPORT_PREFIX_P168E = "P168E_LOCAL_CACHE_BUILD_FROM_VALIDATED_SNAPSHOTS_OR_WAIT_"
CACHE_PACKAGE_DIR = "P168E_LOCAL_CACHE_PACKAGE"
MAX_CACHE_ROWS_PER_SNAPSHOT = 500

SAFETY_FLAGS: dict[str, bool] = {
    "review_only": True,
    "runtime_prompt_modified": False,
    "apply_allowed": False,
    "google_sheets_write": False,
    "live_google_api_call_from_python": False,
    "apps_script_execution": False,
    "clasp_push": False,
    "broker": False,
    "order": False,
    "sizing": False,
    "public_deploy": False,
}

CACHE_PLAN_COLUMNS = [
    "snapshot_id",
    "source_key",
    "tab_name",
    "bounded_range",
    "resolved_path",
    "file_type",
    "cache_action",
    "cache_scope",
    "write_allowed",
    "issue_code",
    "issue_message",
]

CACHE_INDEX_COLUMNS = [
    "snapshot_id",
    "source_key",
    "tab_name",
    "bounded_range",
    "source_file",
    "cache_file",
    "row_count_cached",
    "column_count_cached",
    "checksum_sha256",
    "cache_status",
    "write_allowed",
]

CACHE_VALIDATION_COLUMNS = [
    "snapshot_id",
    "source_file_exists",
    "source_file_type",
    "source_file_readable",
    "cache_file_created",
    "cache_file_exists",
    "row_count_cached",
    "column_count_cached",
    "issue_code",
    "issue_message",
]

WAITING_CHECKLIST_COLUMNS = [
    "snapshot_id",
    "source_key",
    "tab_name",
    "bounded_range",
    "expected_local_file",
    "required_action",
    "operator_note",
]


def utc_now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def utc_stamp() -> str:
    return datetime.now(UTC).strftime("%Y%m%d_%H%M%S")


def discover_latest_export(project_root: Path, prefix: str) -> Path | None:
    exports_root = project_root / "05_EXPORTS"
    if not exports_root.exists():
        return None
    candidates = [
        path for path in exports_root.iterdir() if path.is_dir() and path.name.startswith(prefix)
    ]
    return sorted(candidates, key=lambda path: path.name)[-1] if candidates else None


def discover_latest_p168d_export(project_root: Path) -> Path | None:
    return discover_latest_export(project_root, EXPORT_PREFIX_P168D)


def read_json_if_exists(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"json_error": str(path)}
    return value if isinstance(value, dict) else {"json_not_object": str(path)}


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def safe_snapshot_filename(snapshot_id: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in snapshot_id).strip(
        "_"
    )
    return cleaned or "snapshot"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_csv_snapshot(
    path: Path, max_rows: int = MAX_CACHE_ROWS_PER_SNAPSHOT
) -> tuple[list[str], list[dict[str, str]], int, int]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        records: list[dict[str, str]] = []
        total_rows = 0
        for row in reader:
            total_rows += 1
            if len(records) < max_rows:
                records.append({key: row.get(key, "") for key in fieldnames})
    return fieldnames, records, total_rows, len(fieldnames)


def read_xlsx_snapshot(
    path: Path, max_rows: int = MAX_CACHE_ROWS_PER_SNAPSHOT
) -> tuple[list[str], list[dict[str, str]], int, int, str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:
        return [], [], 0, 0, "XLSX_READER_OPENPYXL_NOT_AVAILABLE"
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_values = next(rows_iter, None)
        if header_values is None:
            workbook.close()
            return [], [], 0, 0, "XLSX_EMPTY"
        headers = [
            str(value) if value is not None else f"column_{index}"
            for index, value in enumerate(header_values, start=1)
        ]
        records: list[dict[str, str]] = []
        total_rows = 0
        for values in rows_iter:
            total_rows += 1
            if len(records) < max_rows:
                records.append(
                    {
                        header: "" if value is None else str(value)
                        for header, value in zip(headers, values, strict=False)
                    }
                )
        workbook.close()
        return headers, records, total_rows, len(headers), "OK"
    except Exception as exc:  # pragma: no cover - malformed workbook defense
        return [], [], 0, 0, f"XLSX_READ_ERROR:{type(exc).__name__}"


def queue_rows_from_p168d(p168d_export_dir: Path) -> list[dict[str, str]]:
    return read_csv_rows(p168d_export_dir / "P168D_READY_IMPORT_QUEUE.csv")


def waiting_rows_from_p168d(p168d_export_dir: Path) -> list[dict[str, str]]:
    validations = read_csv_rows(p168d_export_dir / "P168D_LOCAL_FILE_VALIDATION.csv")
    return [row for row in validations if row.get("import_ready") != "YES"]


def build_cache_plan(
    queue_rows: list[dict[str, str]], waiting_rows: list[dict[str, str]]
) -> list[dict[str, Any]]:
    plan: list[dict[str, Any]] = []
    for row in queue_rows:
        plan.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "resolved_path": row.get("resolved_path", ""),
                "file_type": row.get("file_type", ""),
                "cache_action": "BUILD_LOCAL_JSON_CACHE",
                "cache_scope": f"MAX_{MAX_CACHE_ROWS_PER_SNAPSHOT}_ROWS_PER_SNAPSHOT",
                "write_allowed": "NO",
                "issue_code": "",
                "issue_message": "",
            }
        )
    for row in waiting_rows:
        plan.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "resolved_path": row.get("resolved_path", ""),
                "file_type": row.get("file_type", ""),
                "cache_action": "WAIT_OPERATOR_EXPORT",
                "cache_scope": "NO_CACHE_BUILD_UNTIL_VALIDATED",
                "write_allowed": "NO",
                "issue_code": row.get("issue_code", "MISSING_LOCAL_FILE"),
                "issue_message": row.get(
                    "issue_message", "Operator export is required before local cache build."
                ),
            }
        )
    return plan


def build_waiting_checklist(waiting_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    checklist: list[dict[str, Any]] = []
    for row in waiting_rows:
        checklist.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "expected_local_file": row.get("local_file", ""),
                "required_action": "EXPORT_TAB_TO_LOCAL_CSV_AND_UPDATE_MANIFEST",
                "operator_note": row.get("issue_message", "Waiting for local operator export."),
            }
        )
    return checklist


def cache_one_snapshot(
    row: dict[str, str], cache_dir: Path
) -> tuple[dict[str, Any], dict[str, Any]]:
    snapshot_id = row.get("snapshot_id", "")
    source_path = Path(row.get("resolved_path", ""))
    file_type = (row.get("file_type", "") or source_path.suffix).upper()
    cache_file = cache_dir / f"{safe_snapshot_filename(snapshot_id)}.json"
    validation: dict[str, Any] = {
        "snapshot_id": snapshot_id,
        "source_file_exists": str(source_path.exists() and source_path.is_file()),
        "source_file_type": file_type,
        "source_file_readable": "False",
        "cache_file_created": "False",
        "cache_file_exists": "False",
        "row_count_cached": "",
        "column_count_cached": "",
        "issue_code": "MISSING_SOURCE_FILE",
        "issue_message": "Source file missing at resolved_path.",
    }
    index: dict[str, Any] = {
        "snapshot_id": snapshot_id,
        "source_key": row.get("source_key", ""),
        "tab_name": row.get("tab_name", ""),
        "bounded_range": row.get("bounded_range", ""),
        "source_file": str(source_path),
        "cache_file": "",
        "row_count_cached": "",
        "column_count_cached": "",
        "checksum_sha256": "",
        "cache_status": "NOT_CREATED",
        "write_allowed": "NO",
    }
    if not source_path.exists() or not source_path.is_file():
        return index, validation
    checksum = sha256_file(source_path)
    suffix = source_path.suffix.lower()
    if suffix == ".csv" or "CSV" in file_type:
        headers, records, total_rows, column_count = read_csv_snapshot(source_path)
        read_status = "OK"
    elif suffix in {".xlsx", ".xlsm"} or "XLSX" in file_type:
        headers, records, total_rows, column_count, read_status = read_xlsx_snapshot(source_path)
    else:
        headers = []
        records = []
        total_rows = 0
        column_count = 0
        read_status = "UNSUPPORTED_FILE_TYPE"
    if read_status != "OK":
        validation.update(
            {
                "source_file_readable": "False",
                "issue_code": read_status,
                "issue_message": "Source file exists but could not be read for local cache build.",
            }
        )
        index.update({"checksum_sha256": checksum, "cache_status": read_status})
        return index, validation
    cache_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": "p168e.local_cache.v1",
        "created_at": utc_now_iso(),
        "snapshot_id": snapshot_id,
        "source_key": row.get("source_key", ""),
        "tab_name": row.get("tab_name", ""),
        "bounded_range": row.get("bounded_range", ""),
        "source_file": str(source_path),
        "checksum_sha256": checksum,
        "source_total_data_rows": total_rows,
        "cached_data_rows": len(records),
        "column_count": column_count,
        "headers": headers,
        "records": records,
        "truncated": total_rows > len(records),
        "safety": SAFETY_FLAGS,
    }
    cache_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    validation.update(
        {
            "source_file_readable": "True",
            "cache_file_created": "True",
            "cache_file_exists": str(cache_file.exists()),
            "row_count_cached": str(len(records)),
            "column_count_cached": str(column_count),
            "issue_code": "",
            "issue_message": "",
        }
    )
    index.update(
        {
            "cache_file": str(cache_file),
            "row_count_cached": str(len(records)),
            "column_count_cached": str(column_count),
            "checksum_sha256": checksum,
            "cache_status": "LOCAL_CACHE_CREATED",
        }
    )
    return index, validation


def _markdown_table(rows: list[dict[str, Any]], fields: list[str]) -> str:
    if not rows:
        return "_No rows._\n"
    header = "| " + " | ".join(fields) + " |"
    separator = "| " + " | ".join(["---"] * len(fields)) + " |"
    lines = [header, separator]
    for row in rows:
        lines.append(
            "| " + " | ".join(str(row.get(field, "")).replace("|", "/") for field in fields) + " |"
        )
    return "\n".join(lines) + "\n"


def build_report(
    *,
    p168d_export_dir: Path | None,
    queue_rows: list[dict[str, str]],
    waiting_rows: list[dict[str, str]],
    cache_index_rows: list[dict[str, Any]],
    blocker_count: int,
) -> str:
    lines = [
        "# P168E — Local cache build from validated snapshots or wait",
        "",
        "## Decision",
        "",
        "Build a local JSON cache only from P168D validated local operator exports. If no validated exports are present, wait without blockers.",
        "",
        "## Source",
        "",
        f"- P168D export: `{p168d_export_dir}`",
        f"- Ready snapshot rows: `{len(queue_rows)}`",
        f"- Waiting snapshot rows: `{len(waiting_rows)}`",
        f"- Cache files created: `{len(cache_index_rows)}`",
        f"- Blocker count: `{blocker_count}`",
        "",
        "## Safety",
        "",
        "- Google Sheets write: `False`",
        "- Live Google API call from Python: `False`",
        "- Apps Script execution: `False`",
        "- CLASP push: `False`",
        "- Broker/order/sizing: `False`",
        "",
        "## Cache index preview",
        "",
        _markdown_table(
            cache_index_rows,
            ["snapshot_id", "tab_name", "cache_status", "row_count_cached", "write_allowed"],
        ),
    ]
    return "\n".join(lines)


def build_outputs(project_root: Path) -> Path:
    project_root = project_root.resolve()
    p168d_export_dir = discover_latest_p168d_export(project_root)
    p168d_summary = (
        read_json_if_exists(p168d_export_dir / "P168D_SUMMARY.json") if p168d_export_dir else {}
    )
    queue_rows = queue_rows_from_p168d(p168d_export_dir) if p168d_export_dir else []
    waiting_rows = waiting_rows_from_p168d(p168d_export_dir) if p168d_export_dir else []
    blockers: list[str] = []
    if not p168d_export_dir:
        blockers.append("P168D_EXPORT_MISSING")
    if p168d_summary.get("hierarchy_locked") is not True:
        blockers.append("P168D_HIERARCHY_NOT_LOCKED")
    if p168d_summary.get("google_sheets_write") is True:
        blockers.append("P168D_GOOGLE_SHEETS_WRITE_TRUE")
    if p168d_summary.get("live_google_api_call_from_python") is True:
        blockers.append("P168D_LIVE_PYTHON_GOOGLE_API_TRUE")
    if p168d_summary.get("preferred_path_now") not in {"LOCAL_EXPORT_IMPORT_FIRST", None}:
        blockers.append("P168D_PREFERRED_PATH_NOT_LOCAL_EXPORT_IMPORT_FIRST")
    export_dir = project_root / "05_EXPORTS" / f"{EXPORT_PREFIX_P168E}{utc_stamp()}"
    export_dir.mkdir(parents=True, exist_ok=False)
    cache_dir = export_dir / CACHE_PACKAGE_DIR
    cache_index_rows: list[dict[str, Any]] = []
    cache_validation_rows: list[dict[str, Any]] = []
    for row in queue_rows:
        index, validation = cache_one_snapshot(row, cache_dir)
        cache_index_rows.append(index)
        cache_validation_rows.append(validation)
    plan_rows = build_cache_plan(queue_rows, waiting_rows)
    waiting_checklist = build_waiting_checklist(waiting_rows)
    blocker_count = len(blockers)
    cache_files_created_count = sum(
        1 for row in cache_validation_rows if row.get("cache_file_created") == "True"
    )
    cache_ready_count = len(queue_rows)
    status = STATUS_READY if cache_ready_count else STATUS_WAIT
    write_csv(export_dir / "P168E_LOCAL_CACHE_BUILD_PLAN.csv", plan_rows, CACHE_PLAN_COLUMNS)
    write_csv(export_dir / "P168E_LOCAL_CACHE_INDEX.csv", cache_index_rows, CACHE_INDEX_COLUMNS)
    write_csv(
        export_dir / "P168E_CACHE_FILE_VALIDATION.csv",
        cache_validation_rows,
        CACHE_VALIDATION_COLUMNS,
    )
    write_csv(
        export_dir / "P168E_WAITING_EXPORTS_CHECKLIST.csv",
        waiting_checklist,
        WAITING_CHECKLIST_COLUMNS,
    )
    report = build_report(
        p168d_export_dir=p168d_export_dir,
        queue_rows=queue_rows,
        waiting_rows=waiting_rows,
        cache_index_rows=cache_index_rows,
        blocker_count=blocker_count,
    )
    (export_dir / "P168E_CACHE_BUILD_REPORT.md").write_text(report, encoding="utf-8")
    summary: dict[str, Any] = {
        "step": STEP,
        "status": status,
        "project": "MVP_QAIC_PY",
        "hierarchy_locked": p168d_summary.get("hierarchy_locked") is True,
        "p168d_export_dir": str(p168d_export_dir) if p168d_export_dir else "",
        "preferred_path_now": "LOCAL_EXPORT_IMPORT_FIRST",
        "api_binding_later": "GOOGLE_SHEETS_API_READONLY",
        "ready_snapshot_count": cache_ready_count,
        "waiting_snapshot_count": len(waiting_rows),
        "cache_files_created_count": cache_files_created_count,
        "cache_index_count": len(cache_index_rows),
        "cache_package_dir": str(cache_dir) if cache_files_created_count else "",
        "max_cache_rows_per_snapshot": MAX_CACHE_ROWS_PER_SNAPSHOT,
        "next": "P168F_OPERATOR_EXPORT_CAPTURE_OR_P169_UI_DATA_BINDING_DECISION",
        "blockers": blockers,
        "blocker_count": blocker_count,
        **SAFETY_FLAGS,
    }
    (export_dir / "P168E_SUMMARY.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print("P168E_LOCAL_CACHE_BUILD_READY_READONLY")
    print(f"hierarchy_locked={summary['hierarchy_locked']}")
    print(f"ready_snapshot_count={summary['ready_snapshot_count']}")
    print(f"waiting_snapshot_count={summary['waiting_snapshot_count']}")
    print(f"cache_files_created_count={summary['cache_files_created_count']}")
    print(f"runtime_prompt_modified={summary['runtime_prompt_modified']}")
    print(f"apply_allowed={summary['apply_allowed']}")
    print(f"google_sheets_write={summary['google_sheets_write']}")
    print(f"live_google_api_call_from_python={summary['live_google_api_call_from_python']}")
    print(f"blocker_count={summary['blocker_count']}")
    print(f"output_dir={export_dir}")
    print(f"next={summary['next']}")
    return export_dir


def main() -> None:
    parser = argparse.ArgumentParser(description=STEP)
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    build_outputs(args.project_root)


if __name__ == "__main__":
    main()
