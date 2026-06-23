"""P168D local snapshot import reader from operator exports.

This module stays inside MVP_QAIC_PY. It does not call Google APIs, does not
write Google Sheets, does not execute Apps Script, and does not touch broker
or order flows. It consumes the P168C read-only binding contract and prepares a
safe local import reader for operator-provided CSV/XLSX snapshots.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

STEP = "P168D_LOCAL_SNAPSHOT_IMPORT_READER_FROM_OPERATOR_EXPORTS"
STATUS_READY_WAIT_OPERATOR_EXPORTS = (
    "P168D_LOCAL_SNAPSHOT_IMPORT_READER_READY_WAIT_OPERATOR_EXPORTS_READONLY"
)
STATUS_READY_WITH_FILES = "P168D_LOCAL_SNAPSHOT_IMPORT_READER_READY_WITH_LOCAL_FILES_READONLY"
EXPORT_PREFIX_P168C = "P168C_SHEETS_READONLY_EXPORT_IMPORT_OR_API_BINDING_"
EXPORT_PREFIX_P168D = "P168D_LOCAL_SNAPSHOT_IMPORT_READER_FROM_OPERATOR_EXPORTS_"
DEFAULT_OPERATOR_EXPORT_DIR = Path("01_INPUTS") / "P168D_OPERATOR_SHEETS_EXPORTS"
DEFAULT_OPERATOR_MANIFEST = "P168D_OPERATOR_EXPORT_MANIFEST.csv"

SAFETY_FLAGS: dict[str, bool] = {
    "review_only": True,
    "runtime_prompt_modified": False,
    "apply_allowed": False,
    "google_sheets_write": False,
    "live_google_api_call_from_python": False,
    "apps_script_execution": False,
    "clasp_push": False,
    "broker": False,
    "order": False,
    "sizing": False,
    "public_deploy": False,
}

MANIFEST_COLUMNS = [
    "snapshot_id",
    "source_key",
    "spreadsheet_id",
    "spreadsheet_title",
    "tab_name",
    "bounded_range",
    "local_file",
    "file_type",
    "row_count",
    "column_count",
    "captured_at",
    "checksum_sha256",
    "operator_review_status",
]

VALIDATION_COLUMNS = [
    "snapshot_id",
    "source_key",
    "tab_name",
    "bounded_range",
    "local_file",
    "resolved_path",
    "file_exists",
    "file_type",
    "reader_status",
    "row_count_detected",
    "column_count_detected",
    "checksum_sha256_detected",
    "operator_review_status",
    "import_ready",
    "issue_code",
    "issue_message",
]

IMPORT_QUEUE_COLUMNS = [
    "snapshot_id",
    "source_key",
    "spreadsheet_id",
    "spreadsheet_title",
    "tab_name",
    "bounded_range",
    "local_file",
    "resolved_path",
    "file_type",
    "row_count_detected",
    "column_count_detected",
    "checksum_sha256_detected",
    "import_stage",
    "write_allowed",
]


def utc_now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def utc_stamp() -> str:
    return datetime.now(UTC).strftime("%Y%m%d_%H%M%S")


def discover_latest_export(project_root: Path, prefix: str) -> Path | None:
    exports_root = project_root / "05_EXPORTS"
    if not exports_root.exists():
        return None
    candidates = [p for p in exports_root.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    return sorted(candidates, key=lambda p: p.name)[-1] if candidates else None


def discover_latest_p168c_export(project_root: Path) -> Path | None:
    return discover_latest_export(project_root, EXPORT_PREFIX_P168C)


def read_json_if_exists(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"json_error": str(path)}
    return value if isinstance(value, dict) else {"json_not_object": str(path)}


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def p168c_summary_path(export_dir: Path) -> Path:
    return export_dir / "P168C_SUMMARY.json"


def p168c_manifest_template_path(export_dir: Path) -> Path:
    return export_dir / "P168C_EXPORT_IMPORT_MANIFEST_TEMPLATE.csv"


def normalize_manifest_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    now = utc_now_iso()
    for index, row in enumerate(rows, start=1):
        local_file = row.get("local_file", "").strip()
        snapshot_id = row.get("snapshot_id", "").strip()
        if not snapshot_id or snapshot_id == "PENDING_OPERATOR_EXPORT":
            tab_name = row.get("tab_name", "").strip() or f"row_{index}"
            safe_tab = (
                "".join(ch if ch.isalnum() else "_" for ch in tab_name).strip("_") or f"row_{index}"
            )
            snapshot_id = f"P168D_{safe_tab}_{index:03d}"
        normalized.append(
            {
                "snapshot_id": snapshot_id,
                "source_key": row.get("source_key", ""),
                "spreadsheet_id": row.get("spreadsheet_id", ""),
                "spreadsheet_title": row.get("spreadsheet_title", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "local_file": local_file,
                "file_type": row.get("file_type", "CSV_OR_XLSX") or "CSV_OR_XLSX",
                "row_count": row.get("row_count", "PENDING") or "PENDING",
                "column_count": row.get("column_count", "PENDING") or "PENDING",
                "captured_at": row.get("captured_at", now) or now,
                "checksum_sha256": row.get("checksum_sha256", "PENDING") or "PENDING",
                "operator_review_status": row.get("operator_review_status", "PENDING_EXPORT")
                or "PENDING_EXPORT",
            }
        )
    return normalized


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def detect_csv_shape(path: Path) -> tuple[int, int]:
    row_count = 0
    max_columns = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            row_count += 1
            max_columns = max(max_columns, len(row))
    return row_count, max_columns


def detect_xlsx_shape(path: Path) -> tuple[int, int, str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:
        return 0, 0, "XLSX_READER_OPENPYXL_NOT_AVAILABLE"
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = int(sheet.max_row or 0)
        columns = int(sheet.max_column or 0)
        workbook.close()
        return rows, columns, "OK"
    except Exception as exc:  # pragma: no cover - defensive for malformed workbooks
        return 0, 0, f"XLSX_READ_ERROR:{type(exc).__name__}"


def resolve_local_file(operator_export_dir: Path, local_file: str) -> Path:
    candidate = Path(local_file)
    if candidate.is_absolute():
        return candidate
    return operator_export_dir / candidate


def validate_manifest_rows(
    manifest_rows: list[dict[str, str]],
    operator_export_dir: Path,
) -> list[dict[str, Any]]:
    validations: list[dict[str, Any]] = []
    for row in manifest_rows:
        local_file = row.get("local_file", "").strip()
        resolved = (
            resolve_local_file(operator_export_dir, local_file)
            if local_file
            else operator_export_dir
        )
        exists = bool(local_file) and resolved.exists() and resolved.is_file()
        file_type = (row.get("file_type") or "").upper()
        suffix = resolved.suffix.lower()
        checksum = ""
        detected_rows = ""
        detected_columns = ""
        reader_status = "WAITING_FOR_OPERATOR_EXPORT"
        issue_code = "MISSING_LOCAL_FILE"
        issue_message = (
            "Operator must place the exported snapshot file in the operator export directory."
        )
        import_ready = "NO"
        if exists:
            checksum = sha256_file(resolved)
            if suffix == ".csv" or file_type == "CSV":
                rows, columns = detect_csv_shape(resolved)
                detected_rows = str(rows)
                detected_columns = str(columns)
                reader_status = "CSV_READ_OK"
                issue_code = ""
                issue_message = ""
                import_ready = "YES"
            elif suffix in {".xlsx", ".xlsm"} or "XLSX" in file_type:
                rows, columns, status = detect_xlsx_shape(resolved)
                detected_rows = str(rows) if rows else ""
                detected_columns = str(columns) if columns else ""
                reader_status = status if status != "OK" else "XLSX_READ_OK"
                if status == "OK":
                    issue_code = ""
                    issue_message = ""
                    import_ready = "YES"
                else:
                    issue_code = status
                    issue_message = (
                        "XLSX file exists but cannot be read in the current Python environment."
                    )
            else:
                reader_status = "UNSUPPORTED_FILE_TYPE"
                issue_code = "UNSUPPORTED_FILE_TYPE"
                issue_message = "Use CSV first, or XLSX when openpyxl is available."
        validations.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "local_file": local_file,
                "resolved_path": str(resolved),
                "file_exists": str(exists),
                "file_type": row.get("file_type", ""),
                "reader_status": reader_status,
                "row_count_detected": detected_rows,
                "column_count_detected": detected_columns,
                "checksum_sha256_detected": checksum,
                "operator_review_status": row.get("operator_review_status", ""),
                "import_ready": import_ready,
                "issue_code": issue_code,
                "issue_message": issue_message,
            }
        )
    return validations


def build_ready_import_queue(
    manifest_rows: list[dict[str, str]],
    validation_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    validation_by_id = {str(row.get("snapshot_id", "")): row for row in validation_rows}
    queue: list[dict[str, Any]] = []
    for row in manifest_rows:
        validation = validation_by_id.get(row.get("snapshot_id", ""), {})
        if validation.get("import_ready") != "YES":
            continue
        queue.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "spreadsheet_id": row.get("spreadsheet_id", ""),
                "spreadsheet_title": row.get("spreadsheet_title", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "local_file": row.get("local_file", ""),
                "resolved_path": validation.get("resolved_path", ""),
                "file_type": row.get("file_type", ""),
                "row_count_detected": validation.get("row_count_detected", ""),
                "column_count_detected": validation.get("column_count_detected", ""),
                "checksum_sha256_detected": validation.get("checksum_sha256_detected", ""),
                "import_stage": "READY_FOR_LOCAL_CACHE_IMPORT",
                "write_allowed": "NO",
            }
        )
    return queue


def build_reader_contract_rows(manifest_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in manifest_rows:
        rows.append(
            {
                "snapshot_id": row.get("snapshot_id", ""),
                "source_key": row.get("source_key", ""),
                "tab_name": row.get("tab_name", ""),
                "bounded_range": row.get("bounded_range", ""),
                "accepted_file_types": "CSV_FIRST_XLSX_OPTIONAL",
                "read_mode": "LOCAL_FILE_ONLY",
                "write_allowed": "NO",
                "google_api_call_allowed": "NO",
                "schema_policy": "HEADER_ROW_PLUS_BOUNDED_RANGE_METADATA",
            }
        )
    return rows


def _markdown_table(rows: list[dict[str, Any]], fields: list[str]) -> str:
    if not rows:
        return "_No rows._\n"
    header = "| " + " | ".join(fields) + " |"
    separator = "| " + " | ".join(["---"] * len(fields)) + " |"
    lines = [header, separator]
    for row in rows:
        lines.append(
            "| " + " | ".join(str(row.get(field, "")).replace("|", "/") for field in fields) + " |"
        )
    return "\n".join(lines) + "\n"


def build_report(
    *,
    p168c_export_dir: Path | None,
    operator_export_dir: Path,
    p168c_summary: dict[str, Any],
    manifest_rows: list[dict[str, str]],
    validation_rows: list[dict[str, Any]],
    import_queue_rows: list[dict[str, Any]],
    blocker_count: int,
) -> str:
    lines = [
        "# P168D — Local snapshot import reader from operator exports",
        "",
        "## Decision",
        "",
        "Use local operator exports first. The Python reader validates files, counts rows/columns, computes checksums, and prepares a local import queue. It does not write Google Sheets and does not call Google APIs.",
        "",
        "## Source",
        "",
        f"- P168C export: `{p168c_export_dir}`",
        f"- P168C preferred path now: `{p168c_summary.get('preferred_path_now')}`",
        f"- Operator export directory: `{operator_export_dir}`",
        f"- Required manifest rows: `{len(manifest_rows)}`",
        f"- Ready import rows: `{len(import_queue_rows)}`",
        f"- Blocker count: `{blocker_count}`",
        "",
        "## Safety",
        "",
        "- Google Sheets write: `False`",
        "- Live Google API call from Python: `False`",
        "- Apps Script execution: `False`",
        "- CLASP push: `False`",
        "- Broker/order/sizing: `False`",
        "",
        "## Validation preview",
        "",
        _markdown_table(
            validation_rows,
            [
                "tab_name",
                "local_file",
                "file_exists",
                "reader_status",
                "import_ready",
                "issue_code",
            ],
        ),
    ]
    return "\n".join(lines)


def build_outputs(
    project_root: Path,
    *,
    operator_export_dir: Path | None = None,
    operator_manifest: Path | None = None,
) -> Path:
    project_root = project_root.resolve()
    p168c_export_dir = discover_latest_p168c_export(project_root)
    p168c_summary = (
        read_json_if_exists(p168c_summary_path(p168c_export_dir)) if p168c_export_dir else {}
    )
    manifest_template_rows = (
        read_csv_rows(p168c_manifest_template_path(p168c_export_dir)) if p168c_export_dir else []
    )
    resolved_operator_dir = (
        operator_export_dir or project_root / DEFAULT_OPERATOR_EXPORT_DIR
    ).resolve()
    resolved_operator_manifest = (
        operator_manifest or resolved_operator_dir / DEFAULT_OPERATOR_MANIFEST
    )
    if resolved_operator_manifest.exists():
        raw_manifest_rows = read_csv_rows(resolved_operator_manifest)
        manifest_source = "OPERATOR_MANIFEST"
    else:
        raw_manifest_rows = manifest_template_rows
        manifest_source = "P168C_TEMPLATE_WAIT_OPERATOR_EXPORTS"
    manifest_rows = normalize_manifest_rows(raw_manifest_rows)
    validation_rows = validate_manifest_rows(manifest_rows, resolved_operator_dir)
    import_queue_rows = build_ready_import_queue(manifest_rows, validation_rows)
    reader_contract_rows = build_reader_contract_rows(manifest_rows)
    missing_file_count = sum(1 for row in validation_rows if row.get("file_exists") != "True")
    files_found_count = sum(1 for row in validation_rows if row.get("file_exists") == "True")
    blockers: list[str] = []
    if not p168c_export_dir:
        blockers.append("P168C_EXPORT_MISSING")
    if p168c_summary.get("hierarchy_locked") is not True:
        blockers.append("P168C_HIERARCHY_NOT_LOCKED")
    if p168c_summary.get("preferred_path_now") not in {"LOCAL_EXPORT_IMPORT_FIRST", None}:
        blockers.append("P168C_PREFERRED_PATH_NOT_LOCAL_EXPORT_IMPORT_FIRST")
    if p168c_summary.get("google_sheets_write") is True:
        blockers.append("P168C_GOOGLE_SHEETS_WRITE_TRUE")
    if p168c_summary.get("live_google_api_call_from_python") is True:
        blockers.append("P168C_LIVE_PYTHON_GOOGLE_API_TRUE")
    if not manifest_rows:
        blockers.append("MANIFEST_ROWS_MISSING")
    blocker_count = len(blockers)
    status = STATUS_READY_WITH_FILES if import_queue_rows else STATUS_READY_WAIT_OPERATOR_EXPORTS
    export_dir = project_root / "05_EXPORTS" / f"{EXPORT_PREFIX_P168D}{utc_stamp()}"
    export_dir.mkdir(parents=True, exist_ok=False)
    write_csv(
        export_dir / "P168D_OPERATOR_EXPORT_MANIFEST_TEMPLATE.csv", manifest_rows, MANIFEST_COLUMNS
    )
    write_csv(export_dir / "P168D_IMPORT_READER_CONTRACT.csv", reader_contract_rows)
    write_csv(export_dir / "P168D_LOCAL_FILE_VALIDATION.csv", validation_rows, VALIDATION_COLUMNS)
    write_csv(export_dir / "P168D_READY_IMPORT_QUEUE.csv", import_queue_rows, IMPORT_QUEUE_COLUMNS)
    report = build_report(
        p168c_export_dir=p168c_export_dir,
        operator_export_dir=resolved_operator_dir,
        p168c_summary=p168c_summary,
        manifest_rows=manifest_rows,
        validation_rows=validation_rows,
        import_queue_rows=import_queue_rows,
        blocker_count=blocker_count,
    )
    (export_dir / "P168D_IMPORT_READER_REPORT.md").write_text(report, encoding="utf-8")
    summary: dict[str, Any] = {
        "step": STEP,
        "status": status,
        "project": "MVP_QAIC_PY",
        "hierarchy_locked": p168c_summary.get("hierarchy_locked") is True,
        "p168c_export_dir": str(p168c_export_dir) if p168c_export_dir else "",
        "operator_export_dir": str(resolved_operator_dir),
        "operator_manifest": str(resolved_operator_manifest),
        "manifest_source": manifest_source,
        "manifest_required_count": len(manifest_rows),
        "files_found_count": files_found_count,
        "files_missing_count": missing_file_count,
        "import_ready_count": len(import_queue_rows),
        "reader_contract_count": len(reader_contract_rows),
        "preferred_path_now": "LOCAL_EXPORT_IMPORT_FIRST",
        "api_binding_later": "GOOGLE_SHEETS_API_READONLY",
        "next": "P168E_LOCAL_CACHE_BUILD_FROM_VALIDATED_SNAPSHOTS_OR_WAIT",
        "blockers": blockers,
        "blocker_count": blocker_count,
        **SAFETY_FLAGS,
    }
    (export_dir / "P168D_SUMMARY.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print("P168D_LOCAL_SNAPSHOT_IMPORT_READER_READY_READONLY")
    print(f"hierarchy_locked={summary['hierarchy_locked']}")
    print(f"manifest_required_count={summary['manifest_required_count']}")
    print(f"files_found_count={summary['files_found_count']}")
    print(f"import_ready_count={summary['import_ready_count']}")
    print(f"runtime_prompt_modified={summary['runtime_prompt_modified']}")
    print(f"apply_allowed={summary['apply_allowed']}")
    print(f"google_sheets_write={summary['google_sheets_write']}")
    print(f"blocker_count={summary['blocker_count']}")
    print(f"output_dir={export_dir}")
    print(f"next={summary['next']}")
    return export_dir


def main() -> None:
    parser = argparse.ArgumentParser(description=STEP)
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    parser.add_argument("--operator-export-dir", type=Path, default=None)
    parser.add_argument("--operator-manifest", type=Path, default=None)
    args = parser.parse_args()
    build_outputs(
        args.project_root,
        operator_export_dir=args.operator_export_dir,
        operator_manifest=args.operator_manifest,
    )


if __name__ == "__main__":
    main()
